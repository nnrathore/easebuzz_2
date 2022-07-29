import openpyxl
from openpyxl import Workbook
def loadinput(sheetname):
    wb = openpyxl.load_workbook("C:\\DDFramework\\Common_files\\Inputsheet.xlsx")
    # ws = wb.active
    ws = wb[sheetname]

    r = ws.max_row
    c = ws.max_column
    print(r)
    print(c)
    for i in range (1,c):
        if ws.cell(1,i).value == "Execution Flag":
            exf= i
        if ws.cell(1,i).value == "Test Name":
            tn= i

    print(exf)
    print(tn)

    for j in range(1,r):
        if ws.cell(j,exf).value =="Yes":
            print(ws.cell(j,tn).value)

    wb.save("test.xlsx")
    wb.close()

