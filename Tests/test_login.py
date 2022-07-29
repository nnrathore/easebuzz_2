import time
from Locators import Login
from selenium import webdriver
from allure_commons.types import AttachmentType
from selenium.webdriver.common.by import By
import selenium
import openpyxl
from Methods.selenium_driver import SeleniumDriver
from Methods.webdriverfactory import WebDriverFactory
import ddt


from Methods import selenium_driver
def DDT(sheetname,start_row, End_Row):

    wb = openpyxl.load_workbook(sheetname)
    # ws = wb.active
    ws = wb["Sheet1"]
    for i in range (start_row,End_Row):
        ws.cell(i,4).value = test_1(ws.cell(i,1).value,ws.cell(i,2).value,ws.cell(i,3).value)
        if ws.cell(i,4).value == "Fail":
            return "Fail"

def test_1(LoginID, login_Password, Error):
    baseUrl = "https://courses.letskodeit.com/practice"
    driver = webdriver.Chrome()
    driver.maximize_window()
    driver.get(baseUrl)
    driver.find_element(By.XPATH,Login.login["lodinid"]).send_keys(LoginID)
    driver.find_element(By.XPATH, Login.login["password"]).send_keys(login_Password)
    driver.find_element(By.XPATH, Login.login["submit"]).click()
    time.sleep(5)
    print(LoginID)
    print(login_Password)
    print(Error)
    # if a == Error:
    #     statuss = "Pass"
    # else:
    statuss = "Fail"
    return statuss

