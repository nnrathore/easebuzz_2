import openpyxl
import os



def loadinput(sheetname, browser_name,inputsheet):

    wb = openpyxl.load_workbook(inputsheet)
    # ws = wb.active
    # print("sheet name is :" + sheetname)
    ws = wb[sheetname]

    r = ws.max_row
    c = ws.max_column
    # print(r)
    # print(c)
    for i in range (1,c):
        if ws.cell(1,i).value == "Execution Flag":
            exf= i
        if ws.cell(1,i).value == "Test Name":
            tn= i

    # print(exf)
    # print(tn)

    for j in range(1,r):
        if ws.cell(j,exf).value =="Yes":
            print(ws.cell(j,tn).value)
            print(ws.cell(j,6).value)
            command = "pytest " + ws.cell(j,6).value + " -v -s --alluredir=C:/DDFramework/allure-report/" + browser_name
            print(command)
            os.system(command)

            if (ws.cell(j,9) == "Yess"):
                Inputfile = ws.cell(j,11).value
                Start_Row = ws.cell(j, 12).value
                End_Row = ws.cell(j, 13).value
                ws.cell(j,8).value = DDT(Inputfile,Start_Row,End_Row)
    wb.save("test.xlsx")
    wb.close()

def DDT(sheetname,start_row, End_Row):

    wb = openpyxl.load_workbook(sheetname)
    # ws = wb.active
    ws = wb["Sheet1"]
    for i in range (start_row,End_Row):
        # ws.cell(i,4).value = test_1(ws.cell(i,1).value,ws.cell(i,2).value,ws.cell(i,3).value)
        if ws.cell(i,4).value == "Fail":
            return "Fail"

wb = openpyxl.load_workbook("C:\\DDFramework\\Common_files\\Master.xlsm")
# ws = wb.active
ws = wb["Main"]
url = ws.cell(1,2).value
login_ID =ws.cell(2,2).value
Login_Password =ws.cell(3,2).value
inputsheet = ws.cell(4,2).value

ws = wb["Browsers"]
r = ws.max_row
ws1 = wb["Functional"]
fr= ws1.max_row

for i in range(1,r+1):
    if ws.cell(i, 2).value == "Yes":
        print(ws.cell(i, 1).value)
        browser_name = ws.cell(i, 1).value
        bn=browser_name
        for j in range(1,fr+1):
            if ws1.cell(j, 2).value == "Yes":
                print(ws1.cell(j, 1).value)
                temp = ws1.cell(j, 1).value
                loadinput(temp, browser_name,inputsheet)


os.system("allure serve C:/DDFramework/allure-report/Chrome")
wb.save("test.xlsx")
wb.close()

