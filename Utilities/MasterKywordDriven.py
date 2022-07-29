import openpyxl
import os
"""
File I/O
'w' -> Write-Only Mode
'r' -> Read-only Mode
'r+' -> Read And Write Mode
'a' -> Append Mode
"""


my_file = open("keyword_test.py", "w")
my_file.write("from Methods.selenium_driver import SeleniumDriver\n")
my_file.write("import os,time\n\n\n")
my_file.write("def test_demo():\n")
appender = "    "

wb = openpyxl.load_workbook("C:\\DDFramework\\Tests\\test_day18.xlsx")
ws = wb.active
ws = wb["Sheet1"]
r = ws.max_row
for i in range(2,r+1):
    if (ws.cell(i,1).value != None):
        locator = ws.cell(i,2).value
        Valuee = ws.cell(i,3).value
        Others = ws.cell(i,4).value
        tes = "Nilesh"
        print(tes[0:1])

        if ws.cell(i,1).value == "Navigate":
            browser = "Chrome"
            my_file.write(appender + 'driver = SeleniumDriver.getWebDriverInstanc1e("'+ browser + '", "'  + Valuee + '")\n')
        elif ws.cell(i,1).value == "Element_click":
            my_file.write(appender + 'SeleniumDriver.elementClick(driver,"' + locator + '")\n')
        elif ws.cell(i,1).value == "Sendkeys":
            if Valuee[0:2] == "D:":
                teampVar = Valuee[2:]
                my_file.write(
                    appender + 'SeleniumDriver.sendKeys(driver,'  + teampVar +  ',"' + locator + '")\n')
            else:
                my_file.write(appender + 'SeleniumDriver.sendKeys(driver,' + "'" + Valuee + "'" + ',"' + locator + '")\n')
        elif ws.cell(i,1).value == "Wait":
            my_file.write(appender + 'time.sleep('+ str(Valuee) + ')' + '\n')
        elif ws.cell(i,1).value == "Verify_title":
            my_file.write(appender + 'SeleniumDriver.Verify_title(driver,"' + Valuee + '")\n')
        elif ws.cell(i,1).value == "SelectbyVal":
            my_file.write(appender + 'SeleniumDriver.selectbyVal(driver,"' + locator + '","' + Valuee + '")\n')
        elif ws.cell(i, 1).value == "SelectbyText":
            my_file.write(appender + 'SeleniumDriver.selectbyText(driver,"' + locator + '","' + Valuee + '")\n')
        elif ws.cell(i, 1).value == "SelectbyIndex":
            my_file.write(appender + 'SeleniumDriver.selectbyIndex(driver,"' + locator + '","' + Valuee + '")\n')
        elif ws.cell(i, 1).value == "Readtext":
            if Valuee[0:2] == "D:":
                teampVar = Valuee[2:]
            my_file.write(appender + teampVar + ' = SeleniumDriver.Readtext(driver,"' + locator + '")\n')
        elif ws.cell(i, 1).value == "Getattribute":
            if Valuee[0:2] == "D:":
                teampVar = Valuee[2:]
            my_file.write(appender + teampVar + ' = SeleniumDriver.getAttributee(driver,"' + locator + '","' + Others + '")\n')
        elif ws.cell(i, 1).value == "Element_exists":
            my_file.write(appender + 'SeleniumDriver.elementPresenceCheck(driver,"' + locator + '")\n')
    else:
        break


my_file.close()
os.system(r"pytest C:/DDFramework/Utilities/keyword_test.py::test_demo")