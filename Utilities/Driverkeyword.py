"""
File I/O
'w' -> Write-Only Mode
'r' -> Read-only Mode
'r+' -> Read And Write Mode
'a' -> Append Mode
"""
import openpyxl
import os

inten = "    "



my_file = open("try.py", "w")
my_file.write("import os\n")
my_file.write("from Methods.selenium_driver import SeleniumDriver\n")
my_file.write("def test_demo():\n")
wb = openpyxl.load_workbook("C:\\DDFramework\\Common_files\\test1.xlsx")
    # ws = wb.active
ws = wb["Sheet1"]

r = ws.max_row

for i in range (1,r+1):
    if ws.cell(i,1).value == "Navigate":
        my_file.write(inten + 'driver = SeleniumDriver.getWebDriverInstanc1e("Edge", "https://google.co.in")\n')
    elif ws.cell(i,1).value == "Click":
        my_file.write("Click\n")
    elif ws.cell(i,1).value == "Sendkeys":
        my_file.write("Sendkeys\n")
    elif ws.cell(i,1).value == "Validate":
        my_file.write("Validate\n")
    elif ws.cell(i,1).value == "ReadValue":
        my_file.write("ReadValue\n")
    elif ws.cell(i,1).value == "GetProperties":
        my_file.write("GetProperties\n")

my_file.write('os.system("pytest C:\DDFramework\Tests\demo_class_day15.py::test_demo2")')

my_file.close()